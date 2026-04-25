"""Excel workbook parsing helpers."""

from pathlib import Path
from typing import Any, Dict, List

from app.modules.documents_ingestion.models import ExtractionMetadata, ExtractionResult

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

def parse_excel_workbook(file_path: Path) -> ExtractionResult:
    """Parse spreadsheets into sheet summaries and row records."""

    if load_workbook is None:
        raise RuntimeError("Excel parsing dependency missing. Install openpyxl.")

    workbook = load_workbook(filename=str(file_path), data_only=True)
    sheet_data: Dict[str, Any] = {}
    textual_chunks: List[str] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            sheet_data[sheet.title] = {"headers": [], "rows": []}
            continue

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        records: List[Dict[str, Any]] = []

        for row in rows[1:]:
            if all(cell in (None, "") for cell in row):
                continue
            record = {
                headers[index] or f"column_{index + 1}": row[index]
                for index in range(len(headers))
            }
            records.append(record)

        sheet_data[sheet.title] = {
            "headers": headers,
            "rows": records[:100],
            "row_count": len(records),
        }

        textual_chunks.append(
            f"Sheet: {sheet.title}\nHeaders: {', '.join([header for header in headers if header])}\nRows: {len(records)}"
        )

    return ExtractionResult(
        text="\n\n".join(textual_chunks),
        structured_data={"sheets": sheet_data},
        metadata=ExtractionMetadata(
            parser="excel_parser",
            document_kind="excel",
            sheet_names=workbook.sheetnames,
        ),
    )
